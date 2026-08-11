#!/usr/bin/env python3
"""배민 '배달처리비 다운로드' 엑셀 → rider_daily_fees 적재 (세전 수입).

배민 deliverycenter 협력사관리 > 배달처리비 다운로드 로 받은 암호 엑셀(비번=계정ID)에서
라이더별·영업일별 배달처리비(전달완료 + 라이더무귀책 배달취소)와 본사미션 금액을 집계해
Supabase rider_daily_fees 로 멱등 upsert 한다. 규제 데이터 — 목적달성 후 원본 엑셀 파기.

사용:
  python3 ingest_delivery_fees.py <엑셀경로> [--password gs01310] [--dry-run]
환경변수(둘 중 하나 로드): NEXT_PUBLIC_SUPABASE_URL, SUPABASE_SERVICE_ROLE_KEY
  예) node --env-file 은 파이썬에 안 먹으니, `set -a; source ../../.env.local; set +a` 후 실행.
"""
import sys, os, io, json, argparse, urllib.request
from collections import defaultdict

import msoffcrypto
import openpyxl


def col_index(header, *names):
    """헤더에서 여러 후보명 중 처음 일치하는 컬럼 인덱스."""
    for i, h in enumerate(header):
        if h is not None and str(h).strip() in names:
            return i
    return None


def parse(path, password):
    buf = io.BytesIO()
    with open(path, "rb") as f:
        of = msoffcrypto.OfficeFile(f)
        of.load_key(password=password)
        of.decrypt(buf)
    buf.seek(0)
    wb = openpyxl.load_workbook(buf, read_only=True, data_only=True)

    # (userId, date) -> {"fee": int, "mission": int, "completed": int}
    acc = defaultdict(lambda: {"fee": 0, "mission": 0, "completed": 0})
    details = []  # 배달건별 상세(전 건 — 취소 포함)

    def norm_date(v):
        s = str(v)[:10]
        return f"{s[:4]}-{s[4:6]}-{s[6:8]}" if len(s) == 8 and s.isdigit() else s

    # ── 배달 내역 상세: 배달처리비 + 완료건수 + 건별 상세 ──
    ws = wb["배달 내역 상세"]
    it = ws.iter_rows(values_only=True)
    hdr = list(next(it))
    ui = col_index(hdr, "User ID")
    di = col_index(hdr, "운행일")
    si = col_index(hdr, "배달상태")
    fi = col_index(hdr, "배달처리비")
    gi = col_index(hdr, "라이더귀책여부")
    if None in (ui, di, si, fi, gi):
        sys.exit(f"[배달 내역 상세] 필수 컬럼 누락: {dict(User_ID=ui, 운행일=di, 배달상태=si, 배달처리비=fi, 귀책=gi)}")
    # 상세 컬럼(없으면 -1 → None 저장)
    no_i = col_index(hdr, "배달번호")
    sn_i = col_index(hdr, "가게이름", "가맹점명", "가게명")
    pu_i = col_index(hdr, "픽업완료")
    dl_i = col_index(hdr, "전달완료")
    ds_i = col_index(hdr, "거리")
    b_i = col_index(hdr, "기본단가")
    w_i = col_index(hdr, "기상할증")
    x_i = col_index(hdr, "추가할증")
    p_i = col_index(hdr, "피크할증 등", "피크할증")
    rg_i = col_index(hdr, "지역 할증", "지역할증")
    bk_i = col_index(hdr, "대량 할증", "대량할증")

    def cell(row, i):
        return row[i] if i is not None else None

    for r in it:
        uid = r[ui]
        if uid is None:
            continue
        uid = str(uid).strip()
        date = norm_date(r[di])
        st = str(r[si]).strip()
        fee = r[fi] or 0
        blame = str(r[gi]).strip().upper() if r[gi] is not None else ""
        if st == "전달완료":
            acc[(uid, date)]["fee"] += fee
            acc[(uid, date)]["completed"] += 1
        elif st in ("배달취소", "배달 취소") and blame == "N":
            acc[(uid, date)]["fee"] += fee  # 무귀책 취소는 배달처리비 지급
        # 건별 상세(전 건 — 취소 포함). 배달번호 없으면 skip.
        dno = cell(r, no_i)
        if dno is not None and str(dno).strip():
            details.append({
                "delivery_no": str(dno).strip(),
                "admin_rider_id": uid,
                "snapshot_date": date,
                "status": st,
                "store_name": str(cell(r, sn_i)).strip() if cell(r, sn_i) is not None else None,
                "pickup_at": str(cell(r, pu_i)) if cell(r, pu_i) is not None else None,
                "delivered_at": str(cell(r, dl_i)) if cell(r, dl_i) is not None else None,
                "distance_m": int(round(cell(r, ds_i))) if isinstance(cell(r, ds_i), (int, float)) else None,
                "base_fee": int(round(cell(r, b_i) or 0)),
                "weather_fee": int(round(cell(r, w_i) or 0)),
                "extra_fee": int(round(cell(r, x_i) or 0)),
                "peak_fee": int(round(cell(r, p_i) or 0)),
                "region_fee": int(round(cell(r, rg_i) or 0)),
                "bulk_fee": int(round(cell(r, bk_i) or 0)),
                "fee_krw": int(round(fee)),
                "rider_fault": blame == "Y",
            })

    # ── 본사 미션 금액: 당일 미션 지급액 ──
    if "본사 미션 금액" in wb.sheetnames:
        ws = wb["본사 미션 금액"]
        it = ws.iter_rows(values_only=True)
        hdr = list(next(it))
        mu = col_index(hdr, "라이더 유저 아이디", "라이더유저아이디")
        ma = col_index(hdr, "지급 금액", "지급금액")
        md = col_index(hdr, "미션 시작일", "미션시작일")
        if None not in (mu, ma, md):
            for r in it:
                uid = r[mu]
                if uid is None:
                    continue
                uid = str(uid).strip()
                date = str(r[md])[:10]
                if len(date) == 8 and date.isdigit():
                    date = f"{date[:4]}-{date[4:6]}-{date[6:8]}"
                acc[(uid, date)]["mission"] += r[ma] or 0

    wb.close()
    rows = []
    for (uid, date), v in acc.items():
        rows.append({
            "admin_rider_id": uid,
            "snapshot_date": date,
            "fee_krw": int(round(v["fee"])),
            "mission_krw": int(round(v["mission"])),
            "completed_cnt": v["completed"],
            "source": os.path.basename(path),
        })
    return rows, details


def upsert(rows, url, key, table, on_conflict):
    endpoint = url.rstrip("/") + f"/rest/v1/{table}?on_conflict={on_conflict}"
    headers = {
        "apikey": key, "Authorization": f"Bearer {key}",
        "Content-Type": "application/json", "Prefer": "resolution=merge-duplicates,return=minimal",
    }
    total = 0
    for i in range(0, len(rows), 500):  # 청크 upsert
        chunk = rows[i:i + 500]
        req = urllib.request.Request(endpoint, data=json.dumps(chunk).encode(), headers=headers, method="POST")
        with urllib.request.urlopen(req) as resp:
            if resp.status not in (200, 201, 204):
                sys.exit(f"upsert 실패({table}) status={resp.status}: {resp.read()[:200]}")
        total += len(chunk)
    return total


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("excel")
    ap.add_argument("--password", default="gs01310")
    ap.add_argument("--dry-run", action="store_true")
    args = ap.parse_args()

    rows, details = parse(args.excel, args.password)
    dates = sorted({r["snapshot_date"] for r in rows})
    riders = len({r["admin_rider_id"] for r in rows})
    total_fee = sum(r["fee_krw"] for r in rows)
    total_mission = sum(r["mission_krw"] for r in rows)
    canceled = sum(1 for d in details if d["status"] != "전달완료")
    print(f"파싱: 일별 rows={len(rows)} riders={riders} dates={dates[0]}~{dates[-1]} "
          f"배달처리비합={total_fee:,} 미션합={total_mission:,} | 건별 details={len(details)}(취소 {canceled})")

    if args.dry_run:
        print("(dry-run: 적재 안 함)")
        return
    url = os.environ.get("NEXT_PUBLIC_SUPABASE_URL")
    key = os.environ.get("SUPABASE_SERVICE_ROLE_KEY")
    if not url or not key:
        sys.exit("NEXT_PUBLIC_SUPABASE_URL / SUPABASE_SERVICE_ROLE_KEY 환경변수 필요")
    n = upsert(rows, url, key, "rider_daily_fees", "admin_rider_id,snapshot_date")
    print(f"적재 완료: {n} rows → rider_daily_fees")
    nd = upsert(details, url, key, "delivery_fee_details", "delivery_no")
    print(f"적재 완료: {nd} rows → delivery_fee_details")


if __name__ == "__main__":
    main()
