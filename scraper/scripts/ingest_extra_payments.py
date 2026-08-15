#!/usr/bin/env python3
"""바로고 주차별 정산내역서 xlsx → 주간 정산 데이터 멱등 upsert.

두 가지를 함께 적재한다(같은 파일이 유일 소스):
  1) '추가배달료' 시트 → rider_extra_payments (기상할증 보정 등 사후 소급 지급, 정산>추가지급)
  2) '을지' 시트 F열 → rider_weekly_insurance (시간제보험료, 정산>차감정산 자동값)
배민 배달처리비 daily 파일엔 보험 시트가 없고(2시트뿐) 배민 보험료 메뉴는 이메일 발송
방식이라 즉시 스크래핑 불가 — 주간 정산서가 유일한 자동 소스.

사용:
  set -a; source .env.local; set +a
  python3 scripts/ingest_extra_payments.py <xlsx> [--dry-run]

주차(수~화)는 파일명(YYYYMMDD_YYYYMMDD_...)에서 추출.
환경변수: NEXT_PUBLIC_SUPABASE_URL(또는 SUPABASE_URL), SUPABASE_SERVICE_ROLE_KEY
"""
import argparse
import json
import os
import re
import sys
import urllib.request

import openpyxl


def parse_week_from_filename(path):
    # 바로고 원본 '20260805_...' 또는 grider 다운로드 '20260805~20260811_...' 둘 다 수용.
    m = re.match(r"(\d{8})[_~](\d{8})?", os.path.basename(path))
    if not m:
        sys.exit("파일명에서 주차 추출 실패 — --week-start 로 직접 지정하세요")
    fmt = lambda s: f"{s[:4]}-{s[4:6]}-{s[6:8]}"
    start = fmt(m.group(1))
    end = fmt(m.group(2)) if m.group(2) else None
    return start, end


def parse(path):
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    if "추가배달료" not in wb.sheetnames:
        sys.exit("'추가배달료' 시트 없음 — 정산내역서 형식 확인 필요")
    ws = wb["추가배달료"]
    rows = []
    header_seen = False
    for row in ws.iter_rows(values_only=True):
        if not row or len(row) < 6:
            continue
        # 헤더 행: | 라이더ID | 이름 | 지급금액 | 배달정보 | 사유 (첫 컬럼 공백)
        if row[1] == "라이더ID":
            header_seen = True
            continue
        if not header_seen:
            continue
        rid, name, amt, info, reason = row[1], row[2], row[3], row[4], row[5]
        if rid is None or amt is None:
            continue
        rows.append({
            "admin_rider_id": str(rid).strip(),
            "rider_name": str(name).strip() if name is not None else None,
            "amount_krw": int(amt),
            "delivery_info": str(info).strip() if info is not None else "",
            "reason": str(reason).strip() if reason is not None else "",
        })
    return rows


def parse_insurance(path):
    """을지 시트 F열(시간제보험료) — User ID(2) × amount(10). 헤더 'User ID' 행 이후 데이터."""
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    ws = next((wb[n] for n in wb.sheetnames if n.startswith("을지")), None)
    if ws is None:
        return []
    rows = []
    for row in ws.iter_rows(values_only=True):
        if len(row) <= 10:
            continue
        # 데이터 행: NO(1)=숫자, User ID(2) 존재
        if not isinstance(row[1], (int, float)) or not row[2]:
            continue
        amt = row[10]
        amt = int(amt) if isinstance(amt, (int, float)) else 0
        if amt <= 0:
            continue
        rows.append({
            "admin_rider_id": str(row[2]).strip(),
            "rider_name": str(row[3]).strip() if row[3] is not None else None,
            "amount_krw": amt,
        })
    return rows


def upsert(rows, url, key, table="rider_extra_payments",
           on_conflict="week_start,admin_rider_id,delivery_info,reason"):
    endpoint = url.rstrip("/") + f"/rest/v1/{table}?on_conflict={on_conflict}"
    headers = {
        "apikey": key, "Authorization": f"Bearer {key}",
        "Content-Type": "application/json", "Prefer": "resolution=merge-duplicates,return=minimal",
    }
    total = 0
    for i in range(0, len(rows), 500):
        chunk = rows[i:i + 500]
        req = urllib.request.Request(endpoint, data=json.dumps(chunk).encode(), headers=headers, method="POST")
        with urllib.request.urlopen(req) as resp:
            if resp.status not in (200, 201, 204):
                sys.exit(f"upsert 실패 status={resp.status}: {resp.read()[:200]}")
        total += len(chunk)
    return total


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("excel")
    ap.add_argument("--dry-run", action="store_true")
    ap.add_argument("--week-start", help="주 시작(수) YYYY-MM-DD — 파일명 대신 직접 지정")
    ap.add_argument("--week-end", help="주 끝(화) YYYY-MM-DD")
    args = ap.parse_args()

    if args.week_start:
        week_start = args.week_start
        week_end = args.week_end or args.week_start
    else:
        week_start, week_end = parse_week_from_filename(args.excel)
        if not week_end:
            sys.exit("주 끝 날짜를 파일명에서 못 읽음 — --week-end 지정 필요")
    rows = parse(args.excel)
    ins = parse_insurance(args.excel)
    print(f"파싱: 주차 {week_start}~{week_end}")
    print(f"  [추가지급] rows={len(rows)} riders={len({r['admin_rider_id'] for r in rows})} "
          f"합계={sum(r['amount_krw'] for r in rows):,}원")
    print(f"  [시간제보험료] rows={len(ins)} riders={len({r['admin_rider_id'] for r in ins})} "
          f"합계={sum(r['amount_krw'] for r in ins):,}원")

    if args.dry_run:
        print("(dry-run: 적재 안 함)")
        return

    url = os.environ.get("NEXT_PUBLIC_SUPABASE_URL") or os.environ.get("SUPABASE_URL")
    key = os.environ.get("SUPABASE_SERVICE_ROLE_KEY")
    if not url or not key:
        sys.exit("SUPABASE_URL/SUPABASE_SERVICE_ROLE_KEY 환경변수 필요")
    src = os.path.basename(args.excel)
    extra_payload = [{**r, "week_start": week_start, "week_end": week_end, "source": src} for r in rows]
    n1 = upsert(extra_payload, url, key)
    print(f"추가지급 적재 완료: {n1}행")
    if ins:
        ins_payload = [{**r, "week_start": week_start, "week_end": week_end, "source": src} for r in ins]
        n2 = upsert(ins_payload, url, key, table="rider_weekly_insurance",
                    on_conflict="week_start,admin_rider_id")
        print(f"시간제보험료 적재 완료: {n2}행")


if __name__ == "__main__":
    main()
